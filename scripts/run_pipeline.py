#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
import shutil
import subprocess
import time
import xml.etree.ElementTree as ET
from pathlib import Path
from urllib.parse import quote

import openpyxl
import requests
from openpyxl.styles import Alignment


UA = "Mozilla/5.0 excel-pubmed-fulltext-review/1.0"
DEFAULT_SHEET = "综述引用（按引文顺序）"


def safe_filename(text: object, limit: int = 150) -> str:
    value = "" if text is None else str(text)
    value = re.sub(r"[\\/:*?\"<>|]+", " ", value)
    value = re.sub(r"\s+", " ", value).strip().rstrip(".")
    return (value[:limit].strip() or "untitled")


def is_pdf(path: Path) -> bool:
    try:
        return path.exists() and path.stat().st_size > 10_000 and path.read_bytes()[:5] == b"%PDF-"
    except OSError:
        return False


def parse_pmid(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    match = re.search(r"(?:pubmed\.ncbi\.nlm\.nih\.gov/|PMID[:\s]*)(\d{6,9})", text, re.I)
    if match:
        return match.group(1)
    digits = re.findall(r"\d{6,9}", text)
    return digits[0] if digits else ""


def normalize_seq(value: object, fallback: int) -> object:
    if value is None or str(value).strip() == "":
        return fallback
    try:
        number = float(value)
        return int(number) if number.is_integer() else value
    except Exception:
        return value


def choose_column(headers: list[str], explicit: str | None, candidates: list[str]) -> str | None:
    if explicit:
        if explicit not in headers:
            raise SystemExit(f"Column not found: {explicit}. Available columns: {headers}")
        return explicit
    normalized = {h.strip().lower(): h for h in headers if h}
    for candidate in candidates:
        hit = normalized.get(candidate.lower())
        if hit:
            return hit
    for header in headers:
        low = str(header).lower()
        if any(candidate.lower() in low for candidate in candidates):
            return header
    return None


def read_rows(args: argparse.Namespace) -> tuple[list[dict], dict[str, str | None]]:
    wb = openpyxl.load_workbook(args.input, data_only=True)
    sheet_name = args.sheet or (DEFAULT_SHEET if DEFAULT_SHEET in wb.sheetnames else wb.sheetnames[0])
    if sheet_name not in wb.sheetnames:
        raise SystemExit(f"Sheet not found: {sheet_name}. Available sheets: {wb.sheetnames}")
    ws = wb[sheet_name]
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]

    cols = {
        "seq": choose_column(headers, args.seq_col, ["序号", "编号", "No.", "No", "期刊名称"]),
        "title": choose_column(headers, args.title_col, ["文章标题", "文章题目", "Title", "Article Title"]),
        "pmid": choose_column(headers, args.pmid_col, ["PMID", "PubMed ID"]),
        "pubmed_url": choose_column(headers, args.pubmed_url_col, ["PubMed链接", "PubMed Link", "pubmed", "链接"]),
        "doi": choose_column(headers, args.doi_col, ["DOI"]),
    }
    if not cols["title"]:
        raise SystemExit(f"Could not infer title column. Available columns: {headers}")

    rows: list[dict] = []
    for idx, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        raw = dict(zip(headers, values))
        title = raw.get(cols["title"]) if cols["title"] else None
        if not title:
            continue
        pmid = parse_pmid(raw.get(cols["pmid"])) if cols["pmid"] else ""
        if not pmid and cols["pubmed_url"]:
            pmid = parse_pmid(raw.get(cols["pubmed_url"]))
        seq = normalize_seq(raw.get(cols["seq"]) if cols["seq"] else None, idx)
        rows.append({"_raw": raw, "序号": seq, "文章标题": str(title).strip(), "PMID": pmid})
    return rows, cols


def fetch_pubmed_metadata(pmids: list[str], email: str) -> dict[str, dict]:
    metadata: dict[str, dict] = {}
    clean_pmids = [p for p in dict.fromkeys(pmids) if p]
    for i in range(0, len(clean_pmids), 80):
        chunk = clean_pmids[i : i + 80]
        response = requests.get(
            "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi",
            params={"db": "pubmed", "id": ",".join(chunk), "retmode": "xml", "tool": "excel-pubmed-fulltext-review", "email": email},
            headers={"User-Agent": UA},
            timeout=30,
        )
        response.raise_for_status()
        root = ET.fromstring(response.text)
        for article in root.findall(".//PubmedArticle"):
            pmid = article.findtext(".//MedlineCitation/PMID")
            if not pmid:
                continue
            entry = {"doi": "", "pmcid": "", "abstract": "", "pubmed_title": ""}
            entry["pubmed_title"] = "".join(article.findall(".//ArticleTitle")[0].itertext()).strip() if article.findall(".//ArticleTitle") else ""
            for aid in article.findall(".//PubmedData/ArticleIdList/ArticleId"):
                id_type = (aid.attrib.get("IdType") or "").lower()
                if id_type == "doi":
                    entry["doi"] = (aid.text or "").strip()
                elif id_type == "pmc":
                    entry["pmcid"] = (aid.text or "").strip()
            abstracts = []
            for abstract in article.findall(".//Abstract/AbstractText"):
                label = abstract.attrib.get("Label")
                text = "".join(abstract.itertext()).strip()
                if text:
                    abstracts.append(f"{label}: {text}" if label else text)
            entry["abstract"] = "\n".join(abstracts)
            metadata[pmid] = entry
        time.sleep(0.34)
    return metadata


def download_url(url: str, out: Path) -> tuple[bool, str]:
    tmp = out.with_suffix(out.suffix + ".part")
    try:
        with requests.get(url, headers={"User-Agent": UA}, timeout=60, stream=True, allow_redirects=True) as response:
            if response.status_code >= 400:
                return False, f"HTTP {response.status_code}"
            with tmp.open("wb") as handle:
                for chunk in response.iter_content(chunk_size=1024 * 128):
                    if chunk:
                        handle.write(chunk)
        if not is_pdf(tmp):
            size = tmp.stat().st_size if tmp.exists() else 0
            tmp.unlink(missing_ok=True)
            return False, f"not a valid PDF ({size} bytes)"
        tmp.replace(out)
        return True, url
    except Exception as exc:
        tmp.unlink(missing_ok=True)
        return False, str(exc)


def try_pmc_pdf(pmcid: str, out: Path) -> tuple[bool, str]:
    if not pmcid:
        return False, "no PMCID"
    urls = [
        f"https://pmc.ncbi.nlm.nih.gov/articles/{pmcid}/pdf/",
        f"https://www.ncbi.nlm.nih.gov/pmc/articles/{pmcid}/pdf/",
    ]
    for url in urls:
        ok, message = download_url(url, out)
        if ok:
            return True, message
    return False, "PMC PDF not available"


def try_unpaywall_pdf(doi: str, out: Path, email: str) -> tuple[bool, str]:
    if not doi:
        return False, "no DOI"
    try:
        response = requests.get(
            f"https://api.unpaywall.org/v2/{quote(doi, safe='')}",
            params={"email": email},
            headers={"User-Agent": UA},
            timeout=25,
        )
        if response.status_code >= 400:
            return False, f"HTTP {response.status_code}"
        data = response.json()
        candidates = []
        best = data.get("best_oa_location") or {}
        if best.get("url_for_pdf"):
            candidates.append(best["url_for_pdf"])
        for loc in data.get("oa_locations") or []:
            if loc.get("url_for_pdf") and loc["url_for_pdf"] not in candidates:
                candidates.append(loc["url_for_pdf"])
        for url in candidates[:5]:
            ok, message = download_url(url, out)
            if ok:
                return True, f"Unpaywall: {message}"
        return False, "no usable Unpaywall PDF"
    except Exception as exc:
        return False, str(exc)


def try_scansci_pdf(doi: str, out: Path, tmp_dir: Path) -> tuple[bool, str]:
    if not doi:
        return False, "no DOI"
    try:
        import scansci_pdf.sources as sources
        from scansci_pdf.sources import download

        tmp_dir.mkdir(parents=True, exist_ok=True)
        result = download(
            doi,
            tmp_dir,
            scihub_enabled=False,
            use_tor=False,
            use_vpnsci=False,
            bibtex=False,
            strategy="legal_only",
            rename=False,
        )
        if result.get("success") and result.get("file"):
            src = Path(result["file"])
            if is_pdf(src):
                shutil.move(str(src), str(out))
                return True, result.get("source", "scansci-pdf legal_only")
        return False, result.get("error") or json.dumps(result, ensure_ascii=False)[:400]
    except Exception as exc:
        return False, str(exc)


def extract_text(pdf: Path, txt: Path) -> tuple[int, str]:
    if not is_pdf(pdf):
        return 0, ""
    try:
        subprocess.run(["pdftotext", "-layout", str(pdf), str(txt)], check=False, timeout=120)
        if txt.exists():
            text = txt.read_text(errors="ignore")
            return len(text), text[:5000]
    except Exception:
        pass
    return 0, ""


def get_raw(row: dict, names: list[str]) -> str:
    raw = row.get("_raw", {})
    for name in names:
        value = raw.get(name)
        if value not in (None, ""):
            return str(value).strip()
    return ""


def build_interpretation(row: dict, meta: dict, pdf_path: str, text_len: int) -> str:
    seq = row["序号"]
    title = row["文章标题"]
    purpose = get_raw(row, ["研究目的", "目的", "Objective", "Aim"])
    study_type = get_raw(row, ["研究类型", "Study type", "Design"])
    methods = get_raw(row, ["研究方法", "方法", "Methods"])
    population = get_raw(row, ["研究对象", "对象", "Cohort", "Population"])
    results = get_raw(row, ["主要研究结果", "主要结果", "Results", "Findings"])
    conclusion = get_raw(row, ["研究结论与意义", "结论", "Conclusion", "Significance"])
    novelty = get_raw(row, ["研究亮点或创新点", "创新点", "Novelty"])

    lines = [
        f"# {seq}. {title}",
        "",
        f"- PMID: {row.get('PMID') or '未提供'}",
        f"- DOI: {meta.get('doi') or '未检索到'}",
        f"- PMCID: {meta.get('pmcid') or '未检索到'}",
        f"- 全文PDF: {pdf_path or '未能下载'}",
        f"- 提取文本量: {text_len} 字符",
        "",
        "## 重点解读",
        f"1. 研究目的：{purpose or '原表未提供；需结合全文或摘要进一步提炼。'}",
        f"2. 研究类型：{study_type or '原表未提供；需结合全文或摘要进一步判断。'}",
        f"3. 研究方法：{methods or '原表未提供；建议从全文 Methods 部分补充。'}",
        f"4. 研究对象：{population or '原表未提供；建议核对样本来源、癌种、治疗线数和分组。'}",
        f"5. 主要结果：{results or '原表未提供；建议从全文 Results/figures 提取关键结局。'}",
        f"6. 结论意义：{conclusion or '原表未提供；需结合讨论部分概括临床或机制意义。'}",
        f"7. 创新点：{novelty or '原表未提供；需结合研究设计、技术路线或临床转化价值判断。'}",
        "",
        "## 综述写作提示",
        "将该文放入综述时，优先提取其免疫细胞/分子标志物、肿瘤微环境特征、疗效或预后指标、验证队列与临床转化价值。避免超出原文证据范围。",
    ]
    if meta.get("abstract"):
        lines.extend(["", "## PubMed摘要", meta["abstract"]])
    if text_len < 1000 and pdf_path:
        lines.extend(["", "## 注意", "全文文本提取量偏低，可能是扫描版、图片型 PDF 或下载内容异常，建议人工核对。"])
    return "\n".join(lines).strip() + "\n"


def target_base(seq: object, title: str) -> str:
    if isinstance(seq, int):
        return f"{seq:02d}_{safe_filename(title)}"
    return f"{safe_filename(seq, 30)}_{safe_filename(title)}"


def write_institutional_checklist(failed: list[dict], out: Path) -> None:
    if not failed:
        out.unlink(missing_ok=True)
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "待机构补全文献"
    ws.append(["序号", "PMID", "DOI", "文章题目", "目标PDF文件名", "PDF路径", "错误"])
    for item in failed:
        ws.append([
            item["序号"],
            item["PMID"],
            item["DOI"],
            item["文章标题"],
            Path(item["PDF路径"] or item["目标PDF路径"]).name,
            item["目标PDF路径"],
            item["错误"],
        ])
    widths = [8, 14, 32, 80, 80, 90, 80]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(out)


def run(args: argparse.Namespace) -> None:
    output_dir = args.output_dir.resolve()
    pdf_dir = output_dir / "下载全文"
    reading_dir = output_dir / "文献重点解读"
    text_dir = output_dir / "全文文本提取"
    tmp_dir = output_dir / "_scansci_tmp"
    report_json = output_dir / "文献全文下载报告.json"
    summary_xlsx = output_dir / "全文文献重点解读汇总.xlsx"
    checklist_xlsx = output_dir / "机构全文补全清单.xlsx"

    for directory in [pdf_dir, reading_dir, text_dir, tmp_dir]:
        directory.mkdir(parents=True, exist_ok=True)

    rows, cols = read_rows(args)
    pmids = [row["PMID"] for row in rows if row["PMID"]]
    metadata = fetch_pubmed_metadata(pmids, args.email) if pmids else {}
    report = []

    for index, row in enumerate(rows, start=1):
        seq = row["序号"]
        title = row["文章标题"]
        pmid = row.get("PMID", "")
        raw_doi = row["_raw"].get(cols["doi"]) if cols.get("doi") else ""
        meta = metadata.get(pmid, {}).copy()
        if raw_doi and not meta.get("doi"):
            meta["doi"] = str(raw_doi).strip()
        base = target_base(seq, title)
        pdf_out = pdf_dir / f"{base}.pdf"
        txt_out = text_dir / f"{base}.txt"
        md_out = reading_dir / f"{base}_重点解读.md"

        status = "failed"
        source = ""
        error = ""
        if is_pdf(pdf_out):
            status = "success"
            source = "existing"
        elif not args.no_download:
            ok, message = try_pmc_pdf(meta.get("pmcid", ""), pdf_out)
            if ok:
                status, source = "success", f"PMC: {message}"
            else:
                pmc_error = message
                ok, message = try_unpaywall_pdf(meta.get("doi", ""), pdf_out, args.email)
                if ok:
                    status, source = "success", message
                else:
                    unpaywall_error = message
                    ok, message = try_scansci_pdf(meta.get("doi", ""), pdf_out, tmp_dir) if args.use_scansci else (False, "disabled")
                    if ok:
                        status, source = "success", f"ScanSci: {message}"
                    else:
                        error = f"PMC: {pmc_error}; Unpaywall: {unpaywall_error}; ScanSci: {message}"
                        pdf_out.unlink(missing_ok=True)
        else:
            error = "download disabled"

        text_len, _ = extract_text(pdf_out, txt_out) if status == "success" else (0, "")
        pdf_path = str(pdf_out) if status == "success" else ""
        md_out.write_text(build_interpretation(row, meta, pdf_path, text_len), encoding="utf-8")
        item = {
            "序号": seq,
            "PMID": pmid,
            "文章标题": title,
            "DOI": meta.get("doi", ""),
            "PMCID": meta.get("pmcid", ""),
            "下载状态": status,
            "来源": source,
            "PDF路径": pdf_path,
            "目标PDF路径": str(pdf_out),
            "解读文件": str(md_out),
            "错误": error,
        }
        report.append(item)
        print(f"[{index}/{len(rows)}] {status}: {seq} {title[:70]}")

    report_json.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "重点解读汇总"
    ws.append(["序号", "文章题目", "重点解读内容", "下载状态", "PDF路径", "解读文件", "DOI", "PMCID"])
    for item in report:
        md_text = Path(item["解读文件"]).read_text(encoding="utf-8")
        ws.append([
            item["序号"],
            item["文章标题"],
            md_text,
            item["下载状态"],
            item["PDF路径"],
            item["解读文件"],
            item["DOI"],
            item["PMCID"],
        ])
    widths = [8, 60, 100, 14, 60, 60, 32, 18]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    for row_cells in ws.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(summary_xlsx)

    failed = [item for item in report if item["下载状态"] != "success"]
    write_institutional_checklist(failed, checklist_xlsx)
    successes = len(report) - len(failed)
    print(f"Done. Downloaded {successes}/{len(report)} PDFs.")
    print(f"Summary: {summary_xlsx}")
    if failed:
        print(f"Institutional checklist: {checklist_xlsx}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Download PubMed-linked papers from Excel and summarize them.")
    parser.add_argument("--input", required=True, type=Path, help="Input .xlsx workbook.")
    parser.add_argument("--sheet", default=None, help=f"Worksheet name. Defaults to '{DEFAULT_SHEET}' when present, otherwise first sheet.")
    parser.add_argument("--output-dir", default=Path.cwd(), type=Path, help="Output directory.")
    parser.add_argument("--seq-col", default=None, help="Citation-order/sequence column.")
    parser.add_argument("--title-col", default=None, help="Article title column.")
    parser.add_argument("--pmid-col", default=None, help="PMID column.")
    parser.add_argument("--pubmed-url-col", default=None, help="PubMed URL column.")
    parser.add_argument("--doi-col", default=None, help="DOI column.")
    parser.add_argument("--email", default="research@example.com", help="Email for PubMed/Unpaywall polite API use.")
    parser.add_argument("--no-download", action="store_true", help="Only reuse existing PDFs and generate reports/summaries.")
    parser.add_argument("--use-scansci", action=argparse.BooleanOptionalAction, default=True, help="Enable scansci-pdf legal_only fallback.")
    return parser.parse_args()


if __name__ == "__main__":
    run(parse_args())
